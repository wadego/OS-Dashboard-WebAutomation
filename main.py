# To save time for operators, this web automation will login the dashboard page
# of Ocean Shipping company and do operation staffs.

from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from openpyxl import load_workbook
import configparser

def get_require():
    print("******MENU******")
    print("1: Automatic unloading according to Excel file")
    print("2：Automatic pick-up according to customer number")
    print("3：Check status according to customer number")
    print("******************\n")
    choice = input("Input your choice：")
    return choice

# Read configuration values from config.ini
config = configparser.ConfigParser()
config.read('config.ini')

website = config.get('Website', 'URL')
username = config.get('Credentials', 'username')
password = config.get('Credentials', 'password')
excel_file_path = config.get('FilePath', 'excel_file')

# get user's input
user_input = int(get_require())

# login website of Ocean Shipping Co.
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)
browser = webdriver.Chrome(options=options)

browser.get(website)
browser.find_element(by=By.XPATH, value='//input[@placeholder="请输入用户名"]').send_keys(username)
browser.find_element(by=By.XPATH, value='//input[@placeholder="请输入密码"]').send_keys(password)
browser.find_element(by=By.CSS_SELECTOR, value='.ivu-btn.ivu-btn-primary.ivu-btn-long').click()
sleep(1)

first_level_menu = browser.find_element(by=By.CSS_SELECTOR, value='ul.ivu-menu')
first_level_menu.click()
sleep(1)
second_level_menu = first_level_menu.find_element(by=By.CSS_SELECTOR, value='li.ivu-menu-submenu')
second_level_menu.click()
sleep(1)

if user_input == 1:# Automatic unloading according to Excel file
    # enter "卸货管理" menu
    menu_items = second_level_menu.find_elements(by=By.CSS_SELECTOR, value='li.ivu-menu-item')
    menu_items[1].click()
    sleep(1)
    # input warehouse number
    browser.find_element(by=By.XPATH, value='//input[@placeholder="扫码输入仓库编号"]').send_keys('ono')
    browser.find_elements(by=By.CSS_SELECTOR, value='.ivu-btn.ivu-btn-primary')[1].click()
    sleep(1)
    # open Excel file
    workbook = load_workbook(filename=excel_file_path)
    # get sheet from the file
    sheet = workbook.worksheets[0]
    # read all the rows on the sheet
    for row in sheet.iter_rows():
        # read all the cells on the row
        for cell in row:
            # in-ware number starts with ONO
            cell_str = str(cell.value)
            if cell_str.startswith("ONO"):
                browser.find_element(by=By.XPATH, value='//input[@placeholder="扫码输入入仓号"]').send_keys(cell_str)
                browser.find_elements(by=By.CSS_SELECTOR, value='.ivu-btn.ivu-btn-primary')[2].click()
                sleep(1)
    #needs manual confirmation due to security considerations
    #browser.find_elements(by=By.CSS_SELECTOR, value='.ivu-btn.ivu-btn-primary')[4].click()
elif user_input == 2: # Automatic pick-up according to customer number
    # enter "提货管理" menu
    menu_items = second_level_menu.find_elements(by=By.CSS_SELECTOR, value='li.ivu-menu-item')
    menu_items[2].click()
    sleep(1)
    clientNo = input("请输入客户号：")
    while len(clientNo) != 5:
        clientNo = input("客户号无效,请重新输入：")

    # inquiry the customer number
    browser.find_element(by=By.XPATH, value='//input[@placeholder="客户号"]').send_keys(clientNo)
    browser.find_elements(by=By.CSS_SELECTOR, value='.ivu-btn.ivu-btn-primary')[1].click()
    sleep(1)

    # choose all the packages of the customer
    check_boxs = browser.find_elements(by=By.CSS_SELECTOR, value='.ivu-checkbox-input')

    if check_boxs[0].is_enabled():
        check_boxs[0].click()
        # confirm pick-up
        browser.find_elements(by=By.CSS_SELECTOR, value='.ivu-btn.ivu-btn-primary')[4].click()
        sleep(1)
        # second confirm
        browser.find_elements(by=By.CSS_SELECTOR, value='.ivu-btn.ivu-btn-primary.ivu-btn-large')[1].click()
    else:
        print("No package for this customer, please confirm again")

elif user_input == 3:
    #enter "订单管理" menu
    menu_items = second_level_menu.find_elements(by=By.CSS_SELECTOR, value='li.ivu-menu-item')
    menu_items[0].click()
    sleep(1)
    clientNo = input("Input customer number：")
    while len(clientNo) != 5:
        clientNo = input("Invalid customer number, input again：")

    #define time
    browser.find_elements(by=By.CSS_SELECTOR, value='.ivu-radio-wrapper.ivu-radio-group-item')[1].click()
    browser.find_element(by=By.XPATH, value='//input[@placeholder="客户号"]').send_keys(clientNo)
    browser.find_elements(by=By.CSS_SELECTOR, value='.ivu-btn.ivu-btn-primary')[1].click()
    sleep(1)
    rows = browser.find_elements(by=By.CSS_SELECTOR, value='.ivu-table-row')
    for row in rows:
        columns = browser.find_elements(by=By.CSS_SELECTOR, value='.ivu-table-column-center')
        pay_columns = browser.find_elements(by=By.CSS_SELECTOR, value='.ivu-table-column-isPaid')
        for column in columns:
            print(column.text)

browser.close()